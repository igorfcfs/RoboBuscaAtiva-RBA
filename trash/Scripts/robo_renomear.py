import os
import openpyxl
import re
import subprocess

# Obter o diretório base do usuário atual
diretorio_base_usuario = os.path.expanduser("~")

# Adicionar o subdiretório específico aos arquivos Excel
diretorio_arquivos = os.path.join(diretorio_base_usuario, "Downloads")

# Agora, diretorio_arquivos conterá o caminho para o diretório "Downloads" do usuário atual, independente do nome de usuário.
print("Renomeando os item da pasta Downloads " + diretorio_arquivos)


# Função para extrair a série e o número de uma célula
def extrair_serie_e_numero(celula):
    if not isinstance(celula, str):
        # Se o valor da célula não for uma string, retorna None para série e número
        return None, None

    padrao_serie = r"NM_TURMA é ([^\n]+)"
    padrao_numero = r"CD_ESCOLA é (\d+)"

    serie_encontrada = re.search(padrao_serie, celula)
    numero_encontrado = re.search(padrao_numero, celula)

    serie = serie_encontrada.group(1) if serie_encontrada else None
    numero = numero_encontrado.group(1) if numero_encontrado else None

    return serie, numero

# Função para renomear arquivos Excel com base na série e no número
def renomear_arquivos_excel():
    for pasta_raiz, subpastas, arquivos in os.walk(diretorio_arquivos):
        for arquivo in arquivos:
            if arquivo.endswith(".xlsx"):
                caminho_arquivo = os.path.join(pasta_raiz, arquivo)
                wb = openpyxl.load_workbook(caminho_arquivo)

                for nome_planilha in wb.sheetnames:
                    planilha = wb[nome_planilha]

                    for linha in planilha.iter_rows(min_row=1, max_row=planilha.max_row, min_col=1, max_col=1):
                        for celula in linha:
                            valor_celula = celula.value
                            if valor_celula:
                                serie, numero = extrair_serie_e_numero(valor_celula)
                                if serie and numero:
                                    novo_nome_arquivo = os.path.join(pasta_raiz, f"{serie} - {numero}.xlsx")
                                    os.rename(caminho_arquivo, novo_nome_arquivo)
                                    wb.save(novo_nome_arquivo)
                                    break

# Exemplo de uso
renomear_arquivos_excel()
print("Arquivos renomeados com sucesso")


def executar_script_gerar_planilia_alunos_convocados():
    # Obter o diretório do arquivo em execução
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))

    # Nome do arquivo a ser executado (neste caso, na mesma pasta)
    caminho_segundo_script = "gerar_planilia_alunos_convocados.py"

    # Caminho completo para o segundo script
    caminho_completo = os.path.join(diretorio_atual, caminho_segundo_script)

    # Executar o segundo script
    subprocess.call(["python", caminho_completo])


executar_script_gerar_planilia_alunos_convocados()

